from __future__ import annotations

import base64
import io
import logging
from datetime import date, datetime

from odoo import Command, _, api, fields, models
from odoo.exceptions import UserError, ValidationError
from odoo.tools import float_compare

from odoo.addons.BCA_Seguros.models.res_partner import (
    TIPOS_RED_EXCLUIDOS_POLIZA,
)

from . import plantilla_portafolio

try:
    import openpyxl
except ImportError:  # pragma: no cover - declarado en external_dependencies
    openpyxl = None

_logger = logging.getLogger(__name__)

# Estructura de la plantilla: encabezados en la fila 1, datos desde la fila 2.
# El emparejamiento de columnas es POR NOMBRE de encabezado, no por posición
# (ver _extraer_filas), así que cada aseguradora puede entregar sus columnas en
# cualquier orden mientras respete los títulos en la primera fila.
FILA_ENCABEZADOS = 1
FILA_INICIO_DATOS = 2

# Hojas soportadas → ramo operativo. AUTOS (Qualitas) queda fuera de alcance.
HOJAS_RAMO = {
    'VIDA': 'vida',
    'GMM': 'gmm',
}

# Columnas mínimas que deben existir por hoja para procesarla (subset crítico del
# diccionario; el resto son opcionales/informativas y no abortan la carga).
COLUMNAS_REQUERIDAS = {
    'vida': [
        'Póliza', 'Producto', 'Clave de Agente', 'Nombre del Contratante',
        'Moneda', 'Fecha inicio Vigencia', 'Fecha Fin Vigencia',
        'Frecuencia de Pago',
    ],
    'gmm': [
        'Poliza actual', 'Producto', 'Clave de Agente', 'Nombre del Contratante',
        'Moneda', 'Fecha inicio Vigencia', 'Fecha Fin Vigencia',
        'Frecuencia de Pago',
    ],
}

# Nombre de la columna que contiene el número de póliza, por ramo.
COL_NUMERO_POLIZA = {'vida': 'Póliza', 'gmm': 'Poliza actual'}

# Hoja de beneficiarios (formato largo, una fila por persona) — B05. Es
# independiente de las hojas de póliza: puede venir sola (para pólizas ya
# cargadas) o junto a VIDA/GMM en el mismo archivo. Se referencia contra la
# póliza por su folio (COL_BENEF_POLIZA) + la aseguradora del wizard.
HOJA_BENEFICIARIOS = 'BENEFICIARIOS'
COL_BENEF_POLIZA = 'Póliza'
COL_BENEF_NOMBRE = 'Nombre del Beneficiario'
COL_BENEF_PARENTESCO = 'Parentesco'
COL_BENEF_PORCENTAJE = '% al que tiene Derecho'
COL_BENEF_FECHA_NAC = 'Fecha de Nacimiento'
COLUMNAS_REQUERIDAS_BENEFICIARIOS = [COL_BENEF_POLIZA, COL_BENEF_NOMBRE]

# Hoja de coberturas adicionales (formato largo, una fila por cobertura). Igual
# que beneficiarios: puede venir sola (sobre pólizas ya cargadas) o junto a
# VIDA/GMM. Se referencia contra la póliza por su folio (tolerante a ceros a la
# izquierda) + la aseguradora del wizard. El mapeo a la cobertura del catálogo
# nativo (PTAV) se hace por la DESCRIPCIÓN; el CÓDIGO se conserva como nota.
HOJA_COBERTURAS = 'COBERTURAS'
COL_COB_POLIZA = 'Póliza'
COL_COB_CODIGO = 'Cobertura Adicional'
COL_COB_DESCRIPCION = 'Descripción Plan Suplementario'
COLUMNAS_REQUERIDAS_COBERTURAS = [COL_COB_POLIZA, COL_COB_DESCRIPCION]

PERIODICIDAD_MAP = {
    'mensual': 'mensual',
    'trimestral': 'trimestral',
    'semestral': 'semestral',
    'anual': 'anual',
}
ESTADO_POLIZA_MAP = {
    'vigente': 'activa',
    'activa': 'activa',
    'en vigor': 'activa',
    'vencida': 'vencida',
    'expirada': 'vencida',
    'cancelada': 'cancelada',
    'cancelado': 'cancelada',
}
ESTADO_CIVIL_MAP = {
    'soltero': 'soltero', 'soltera': 'soltero',
    'casado': 'casado', 'casada': 'casado',
    'divorciado': 'divorciado', 'divorciada': 'divorciado',
    'viudo': 'viudo', 'viuda': 'viudo',
    'union libre': 'union_libre', 'unión libre': 'union_libre',
}
GENERO_MAP = {
    'masculino': 'masculino', 'm': 'masculino', 'hombre': 'masculino',
    'femenino': 'femenino', 'f': 'femenino', 'mujer': 'femenino',
}
PARENTESCO_MAP = {
    'conyuge': 'conyuge', 'cónyuge': 'conyuge', 'esposo': 'conyuge',
    'esposa': 'conyuge',
    'hijo': 'hijo', 'hija': 'hijo', 'hijo(a)': 'hijo',
    'padre': 'padre', 'papa': 'padre', 'papá': 'padre',
    'madre': 'madre', 'mama': 'madre', 'mamá': 'madre',
    'hermano': 'hermano', 'hermana': 'hermano', 'hermano(a)': 'hermano',
}


class BcaWizardCargaPortafolio(models.TransientModel):
    _name = 'bca.wizard.carga.portafolio'
    _description = 'Wizard Carga Masiva de Portafolio'

    # No 'required' a nivel de campo: el botón "Descargar plantilla" es un
    # type="object" que guarda el wizard antes de ejecutarse, y un archivo
    # obligatorio bloquearía la descarga (que precisamente sirve para obtener
    # el archivo). La obligatoriedad se valida en _abrir_workbook al validar/grabar.
    archivo: bytes = fields.Binary(string='Archivo de Portafolio (.xlsx)')
    nombre_archivo: str = fields.Char(string='Nombre del Archivo')
    aseguradora_id: int = fields.Many2one(
        'res.partner',
        string='Aseguradora',
        required=True,
        domain=[('bca_tipo', '=', 'aseguradora')],
        default=lambda self: self.env.ref(
            'BCA_Seguros.partner_metlife', raise_if_not_found=False
        ),
    )
    modo: str = fields.Selection(
        [
            ('crear_actualizar', 'Crear y actualizar'),
            ('solo_crear', 'Solo crear (rechazar duplicados)'),
        ],
        string='Modo',
        required=True,
        default='crear_actualizar',
    )
    state: str = fields.Selection(
        [('cargar', 'Cargar'), ('validado', 'Validado')],
        string='Estado',
        default='cargar',
    )
    reporte_html: str = fields.Html(string='Reporte', readonly=True)
    total_filas: int = fields.Integer(string='Total de Filas', readonly=True)
    creadas: int = fields.Integer(string='Creadas', readonly=True)
    actualizadas: int = fields.Integer(string='Actualizadas', readonly=True)
    rechazadas: int = fields.Integer(string='Rechazadas', readonly=True)

    # ------------------------------------------------------------------ #
    # Acciones (dos fases M1)
    # ------------------------------------------------------------------ #
    def action_validar(self) -> dict:
        """Fase 1 (M1): lee y valida el archivo SIN tocar la base de datos."""
        self.ensure_one()
        wb = self._abrir_workbook()
        resultados, total = self._recorrer(wb, dry_run=True)
        self.write({
            'state': 'validado',
            'total_filas': total,
            'reporte_html': self._render_reporte(resultados, fase='validación'),
            'creadas': 0,
            'actualizadas': 0,
            'rechazadas': sum(1 for r in resultados if r['motivo']),
        })
        return self._reabrir()

    def action_grabar(self) -> dict:
        """Fase 2 (M1): graba cada póliza en un savepoint independiente."""
        self.ensure_one()
        if self.state != 'validado':
            raise UserError(_('Primero debe validar el archivo.'))
        inicio = datetime.now()
        wb = self._abrir_workbook()
        resultados, total = self._recorrer(wb, dry_run=False)
        creadas = sum(1 for r in resultados if r['accion'] == 'creada')
        actualizadas = sum(1 for r in resultados if r['accion'] == 'actualizada')
        rechazadas = sum(1 for r in resultados if r['motivo'])
        _logger.info(
            'Carga de portafolio: %s filas en %.2fs (%s creadas, %s actualizadas, '
            '%s rechazadas)',
            total, (datetime.now() - inicio).total_seconds(),
            creadas, actualizadas, rechazadas,
        )
        self.write({
            'total_filas': total,
            'creadas': creadas,
            'actualizadas': actualizadas,
            'rechazadas': rechazadas,
            'reporte_html': self._render_reporte(resultados, fase='grabado'),
        })
        return self._reabrir()

    def action_descargar_plantilla(self) -> dict:
        """Genera la plantilla .xlsx al vuelo y la entrega como descarga.

        El catálogo de columnas vive en ``plantilla_portafolio`` (única fuente
        de verdad, compartida con el script de dev). El adjunto se ata al
        registro transient para que el vacuum de transients lo purgue.
        """
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_('La librería openpyxl no está instalada.'))
        datos = plantilla_portafolio.construir_workbook_bytes()
        adjunto = self.env['ir.attachment'].create({
            'name': 'plantilla_portafolio_BCA.xlsx',
            'datas': base64.b64encode(datos),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': (
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % adjunto.id,
            'target': 'download',
        }

    # ------------------------------------------------------------------ #
    # Lectura del archivo
    # ------------------------------------------------------------------ #
    def _abrir_workbook(self):
        if openpyxl is None:
            raise UserError(_('La librería openpyxl no está instalada.'))
        if not self.archivo:
            raise UserError(_('Debe adjuntar un archivo.'))
        try:
            datos = base64.b64decode(self.archivo)
            return openpyxl.load_workbook(
                io.BytesIO(datos), read_only=True, data_only=True
            )
        except Exception as exc:  # noqa: BLE001
            raise UserError(_('No se pudo leer el archivo Excel: %s') % exc) from exc

    def _recorrer(self, wb, dry_run: bool) -> tuple:
        """Recorre las hojas soportadas y devuelve (resultados, total_filas).

        Cada resultado: {hoja, fila, poliza, accion, motivo}. `accion` es
        'creada'/'actualizada'/'' y `motivo` el rechazo (vacío si OK).
        """
        resultados: list[dict] = []
        total = 0
        hojas_presentes = set(wb.sheetnames)
        # Estructura: al menos una hoja soportada (póliza, beneficiarios o
        # coberturas) debe existir.
        soportadas = [h for h in HOJAS_RAMO if h in hojas_presentes]
        tiene_beneficiarios = HOJA_BENEFICIARIOS in hojas_presentes
        tiene_coberturas = HOJA_COBERTURAS in hojas_presentes
        if not soportadas and not tiene_beneficiarios and not tiene_coberturas:
            raise UserError(_(
                'El archivo no contiene ninguna hoja soportada (%s). Hojas '
                'encontradas: %s'
            ) % (', '.join(list(HOJAS_RAMO)
                           + [HOJA_BENEFICIARIOS, HOJA_COBERTURAS]),
                 ', '.join(wb.sheetnames)))
        # AUTOS u otras hojas no reconocidas: se reportan como fuera de alcance.
        reconocidas = set(HOJAS_RAMO) | {HOJA_BENEFICIARIOS, HOJA_COBERTURAS}
        for hoja in hojas_presentes - reconocidas:
            if hoja.strip().upper() == 'AUTOS':
                resultados.append({
                    'hoja': hoja, 'fila': '-', 'poliza': '-', 'accion': '',
                    'motivo': _('Ramo AUTOS no soportado (omitido).'),
                })

        # 1) Pólizas. Se registran los folios vistos (con su ramo) para que la
        #    hoja de beneficiarios pueda validar en fase VALIDAR pólizas que aún
        #    no existen pero se crearán en esta misma corrida.
        folios_en_archivo: dict[str, str] = {}
        for hoja in soportadas:
            ramo = HOJAS_RAMO[hoja]
            ws = wb[hoja]
            encabezados, filas = self._extraer_filas(ws)
            faltantes = [c for c in COLUMNAS_REQUERIDAS[ramo] if c not in encabezados]
            if faltantes:
                # Error estructural: detiene TODO sin grabar (M1, R-COB-09 análogo).
                raise UserError(_(
                    'La hoja "%s" no tiene las columnas requeridas: %s'
                ) % (hoja, ', '.join(faltantes)))
            for numero_fila, raw in filas:
                total += 1
                folio = self._txt(raw.get(COL_NUMERO_POLIZA[ramo]))
                if folio:
                    folios_en_archivo[folio] = ramo
                resultados.append(self._procesar_fila(hoja, ramo, numero_fila, raw, dry_run))

        # 2) Beneficiarios (después de las pólizas, para que las creadas en esta
        #    corrida ya existan al grabar).
        if tiene_beneficiarios:
            ws = wb[HOJA_BENEFICIARIOS]
            encabezados, filas = self._extraer_filas(ws)
            faltantes = [c for c in COLUMNAS_REQUERIDAS_BENEFICIARIOS
                         if c not in encabezados]
            if faltantes:
                raise UserError(_(
                    'La hoja "%s" no tiene las columnas requeridas: %s'
                ) % (HOJA_BENEFICIARIOS, ', '.join(faltantes)))
            res_benef, n_benef = self._procesar_beneficiarios(
                filas, dry_run, folios_en_archivo)
            resultados.extend(res_benef)
            total += n_benef

        # 3) Coberturas adicionales (después de las pólizas, para que las creadas
        #    en esta corrida ya existan al grabar).
        if tiene_coberturas:
            ws = wb[HOJA_COBERTURAS]
            encabezados, filas = self._extraer_filas(ws)
            faltantes = [c for c in COLUMNAS_REQUERIDAS_COBERTURAS
                         if c not in encabezados]
            if faltantes:
                raise UserError(_(
                    'La hoja "%s" no tiene las columnas requeridas: %s'
                ) % (HOJA_COBERTURAS, ', '.join(faltantes)))
            res_cob, n_cob = self._procesar_coberturas(
                filas, dry_run, folios_en_archivo)
            resultados.extend(res_cob)
            total += n_cob
        return resultados, total

    def _extraer_filas(self, ws) -> tuple:
        """Devuelve (encabezados, [(numero_fila_excel, {col: valor}), ...])."""
        encabezados: list[str] = []
        filas: list[tuple] = []
        for idx, fila_valores in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == FILA_ENCABEZADOS:
                encabezados = [str(v).strip() if v is not None else '' for v in fila_valores]
            elif idx >= FILA_INICIO_DATOS:
                if not any(v not in (None, '') for v in fila_valores):
                    continue  # fila totalmente vacía
                raw = {
                    encabezados[i]: fila_valores[i]
                    for i in range(min(len(encabezados), len(fila_valores)))
                    if encabezados[i]
                }
                filas.append((idx, raw))
        return encabezados, filas

    # ------------------------------------------------------------------ #
    # Procesamiento de la hoja BENEFICIARIOS (formato largo — B05)
    # ------------------------------------------------------------------ #
    def _procesar_beneficiarios(self, filas: list, dry_run: bool,
                                folios_en_archivo: dict) -> tuple:
        """Agrupa las filas por folio de póliza y procesa un grupo por póliza.

        Devuelve (resultados, total_filas). El reemplazo y la validación del
        100%% son POR PÓLIZA (no por fila), así que se agrupa primero.
        """
        grupos: dict[str, list] = {}
        total = 0
        for numero_fila, raw in filas:
            total += 1
            folio = self._txt(raw.get(COL_BENEF_POLIZA))
            grupos.setdefault(folio, []).append((numero_fila, raw))
        resultados = [
            self._procesar_grupo_beneficiarios(folio, items, dry_run, folios_en_archivo)
            for folio, items in grupos.items()
        ]
        return resultados, total

    def _procesar_grupo_beneficiarios(self, folio: str, items: list, dry_run: bool,
                                      folios_en_archivo: dict) -> dict:
        """Procesa (o valida) todos los beneficiarios de UNA póliza.

        REEMPLAZA: borra los beneficiarios existentes y recrea desde el archivo
        (idempotente en re-ejecuciones). En Vida valida que los porcentajes sumen
        100%. Aísla la escritura en un savepoint para no arrastrar otras pólizas.
        """
        filas_num = [i[0] for i in items]
        rango = ('%s–%s' % (filas_num[0], filas_num[-1])
                 if len(filas_num) > 1 else str(filas_num[0]))
        resultado = {
            'hoja': HOJA_BENEFICIARIOS, 'fila': rango, 'poliza': folio or '-',
            'accion': '', 'motivo': '',
        }
        try:
            if not folio:
                raise UserError(_('Falta el folio de la póliza del beneficiario.'))
            poliza = self._resolver_poliza(folio)
            # Ramo: de la póliza existente o, en VALIDAR, de la hoja de póliza
            # del mismo archivo (aún no grabada).
            ramo = poliza.ramo if poliza else folios_en_archivo.get(folio)
            if not poliza and ramo is None:
                raise UserError(_(
                    'No existe la póliza "%s" para esta aseguradora.') % folio)

            beneficiarios = []
            for _numero_fila, raw in items:
                nombre = self._txt(raw.get(COL_BENEF_NOMBRE))
                if not nombre:
                    continue
                beneficiarios.append({
                    'nombre': nombre,
                    'parentesco': self._map_simple(
                        raw.get(COL_BENEF_PARENTESCO), PARENTESCO_MAP),
                    'porcentaje': self._norm_monto(raw.get(COL_BENEF_PORCENTAJE)),
                    'fecha_nacimiento': self._norm_fecha(raw.get(COL_BENEF_FECHA_NAC)),
                })
            if not beneficiarios:
                raise UserError(_(
                    'La póliza "%s" no tiene beneficiarios con nombre válido.') % folio)

            # Regla del 100% solo para Vida (en GMM el mismo modelo son
            # dependientes, sin porcentaje de reparto).
            if ramo == 'vida':
                total_pct = sum(b['porcentaje'] for b in beneficiarios)
                if float_compare(total_pct, 100.0, precision_digits=2) != 0:
                    raise UserError(_(
                        'Los porcentajes de los beneficiarios de la póliza "%s" '
                        'deben sumar 100%% (actual: %.2f%%).') % (folio, total_pct))

            if dry_run:
                resultado['accion'] = 'beneficiarios'
                return resultado

            # En grabar la póliza YA debe existir (las hojas de póliza se
            # procesan antes). Si no, es que ni existía ni se pudo crear.
            if not poliza:
                poliza = self._resolver_poliza(folio)
            if not poliza:
                raise UserError(_(
                    'No existe la póliza "%s" (no se creó en esta carga).') % folio)

            with self.env.cr.savepoint():
                poliza.beneficiario_ids.unlink()  # reemplazo idempotente
                for b in beneficiarios:
                    partner = self._find_or_create_partner({'name': b['nombre']})
                    self.env['bca.poliza.beneficiario'].create({
                        'poliza_id': poliza.id,
                        'beneficiario_id': partner.id,
                        'parentesco': b['parentesco'],
                        'porcentaje': b['porcentaje'],
                        'fecha_nacimiento': b['fecha_nacimiento'],
                    })
            resultado['accion'] = 'beneficiarios'
        except (UserError, ValidationError) as exc:
            resultado['motivo'] = exc.args[0] if exc.args else str(exc)
        except Exception as exc:  # noqa: BLE001
            _logger.exception('Error procesando beneficiarios de la póliza %s', folio)
            resultado['motivo'] = _('Error inesperado: %s') % exc
        return resultado

    # ------------------------------------------------------------------ #
    # Procesamiento de la hoja COBERTURAS (formato largo)
    # ------------------------------------------------------------------ #
    def _procesar_coberturas(self, filas: list, dry_run: bool,
                             folios_en_archivo: dict) -> tuple:
        """Agrupa las filas por folio y procesa un grupo por póliza (reemplazo).

        Devuelve (resultados, total_filas). El índice de pólizas por folio
        (tolerante a ceros a la izquierda) se construye una sola vez.
        """
        grupos: dict[str, list] = {}
        total = 0
        for numero_fila, raw in filas:
            total += 1
            folio = self._txt(raw.get(COL_COB_POLIZA))
            grupos.setdefault(folio, []).append((numero_fila, raw))
        indice = self._indice_polizas_por_folio()
        resultados = [
            self._procesar_grupo_coberturas(
                folio, items, dry_run, indice, folios_en_archivo)
            for folio, items in grupos.items()
        ]
        return resultados, total

    def _procesar_grupo_coberturas(self, folio: str, items: list, dry_run: bool,
                                   indice: dict, folios_en_archivo: dict) -> dict:
        """Procesa (o valida) todas las coberturas adicionales de UNA póliza.

        REEMPLAZA `cobertura_adicional_ids` con las coberturas mapeadas del
        archivo que OFRECE el producto de la póliza (dominio nativo por PTAV), y
        vuelca en las notas (`coberturas_adicionales`) un resumen de TODAS las
        filas —incluidas las no asignables— para no perder dato. Las coberturas
        cuyo producto no las ofrece, sin catálogo, o de una póliza inexistente,
        se omiten (no abortan) y se reportan.
        """
        filas_num = [i[0] for i in items]
        rango = ('%s–%s' % (filas_num[0], filas_num[-1])
                 if len(filas_num) > 1 else str(filas_num[0]))
        resultado = {
            'hoja': HOJA_COBERTURAS, 'fila': rango, 'poliza': folio or '-',
            'accion': '', 'motivo': '',
        }
        try:
            if not folio:
                raise UserError(_('Falta el folio de la póliza de la cobertura.'))
            clave = plantilla_portafolio.normalizar_folio(folio)
            poliza = indice.get(clave)
            en_archivo = clave in {
                plantilla_portafolio.normalizar_folio(f) for f in folios_en_archivo
            }
            if not poliza and not en_archivo:
                # Este archivo NO crea cartera: si la póliza no existe, se omite.
                resultado['motivo'] = _(
                    'No existe la póliza "%s" para esta aseguradora '
                    '(cobertura omitida).') % folio
                return resultado

            attr = self.env.ref(
                'BCA_Seguros.attr_cobertura_adicional', raise_if_not_found=False)
            ptav_ids: set[int] = set()
            lineas_nota: list[str] = []
            asignadas = 0
            omitidas = 0
            for _numero_fila, raw in items:
                codigo = self._txt(raw.get(COL_COB_CODIGO))
                desc = self._txt(raw.get(COL_COB_DESCRIPCION))
                xmlid = plantilla_portafolio.mapear_cobertura(desc)
                estado = ''
                if not xmlid:
                    estado = 'sin catálogo'
                else:
                    ptav = self._resolver_ptav_cobertura(poliza, attr, xmlid)
                    if ptav:
                        ptav_ids.add(ptav.id)
                        asignadas += 1
                    else:
                        estado = 'no ofrecida por el producto'
                if estado:
                    omitidas += 1
                etiqueta = ' — '.join(p for p in (codigo, desc) if p) or '(sin datos)'
                lineas_nota.append(
                    '· %s%s' % (('[%s] ' % estado) if estado else '', etiqueta))

            if dry_run:
                resultado['accion'] = 'coberturas'
                if omitidas and not asignadas:
                    resultado['motivo'] = _(
                        '%s cobertura(s) sin asignación estructurada '
                        '(se guardarán como nota).') % omitidas
                return resultado

            # Fase grabar: la póliza YA debe existir (las hojas de póliza se
            # procesan antes). Si sigue sin existir, se omite.
            if not poliza:
                resultado['motivo'] = _(
                    'No existe la póliza "%s" (no se creó en esta carga; '
                    'cobertura omitida).') % folio
                return resultado

            nota = _('Coberturas MetLife (importadas):') + '\n' + '\n'.join(lineas_nota)
            with self.env.cr.savepoint():
                poliza.cobertura_adicional_ids = [Command.set(list(ptav_ids))]
                poliza.coberturas_adicionales = nota
            resultado['accion'] = 'coberturas'
            if omitidas:
                resultado['motivo'] = _(
                    '%s cobertura(s) sin asignación estructurada '
                    '(guardadas como nota).') % omitidas
        except (UserError, ValidationError) as exc:
            resultado['motivo'] = exc.args[0] if exc.args else str(exc)
        except Exception as exc:  # noqa: BLE001
            _logger.exception('Error procesando coberturas de la póliza %s', folio)
            resultado['motivo'] = _('Error inesperado: %s') % exc
        return resultado

    def _resolver_ptav_cobertura(self, poliza, attr, xmlid: str):
        """Devuelve la PTAV (product.template.attribute.value) de la cobertura
        `xmlid` para el producto de la póliza, o un recordset vacío si el
        producto no la ofrece / no se resuelve. Sólo lectura."""
        Ptav = self.env['product.template.attribute.value']
        if not poliza or not poliza.producto_id or not attr:
            return Ptav
        value = self.env.ref(
            'BCA_Seguros.%s' % xmlid, raise_if_not_found=False)
        if not value:
            return Ptav
        return Ptav.search([
            ('product_tmpl_id', '=', poliza.producto_id.id),
            ('product_attribute_value_id', '=', value.id),
            ('attribute_id', '=', attr.id),
            ('ptav_active', '=', True),
        ], limit=1)

    def _indice_polizas_por_folio(self) -> dict:
        """Índice {folio_normalizado: poliza} de la aseguradora del wizard, para
        emparejar folios del archivo ignorando ceros a la izquierda (el archivo
        de MetLife trae folios tipo '0008312115')."""
        indice: dict = {}
        polizas = self.env['bca.poliza'].search([
            ('aseguradora_id', '=', self.aseguradora_id.id),
        ])
        for pol in polizas:
            indice[plantilla_portafolio.normalizar_folio(pol.name)] = pol
        return indice

    # ------------------------------------------------------------------ #
    # Procesamiento por fila
    # ------------------------------------------------------------------ #
    def _procesar_fila(self, hoja: str, ramo: str, numero_fila: int, raw: dict,
                        dry_run: bool) -> dict:
        numero_poliza = self._txt(raw.get(COL_NUMERO_POLIZA[ramo]))
        resultado = {
            'hoja': hoja, 'fila': numero_fila, 'poliza': numero_poliza or '-',
            'accion': '', 'motivo': '',
        }
        try:
            vals, beneficiarios = self._construir_vals(ramo, raw)
            poliza_existente = self.env['bca.poliza'].search([
                ('name', '=', vals['name']),
                ('aseguradora_id', '=', self.aseguradora_id.id),
            ], limit=1)
            if poliza_existente and self.modo == 'solo_crear':
                resultado['motivo'] = _('Ya existe (modo solo crear).')
                return resultado
            if dry_run:
                resultado['accion'] = 'actualizada' if poliza_existente else 'creada'
                return resultado
            # Fase grabar: aislar cada póliza en su propio savepoint (M1).
            with self.env.cr.savepoint():
                if poliza_existente:
                    self._actualizar_poliza(poliza_existente, vals)
                    resultado['accion'] = 'actualizada'
                else:
                    self._crear_poliza(vals, beneficiarios, ramo)
                    resultado['accion'] = 'creada'
        except UserError as exc:
            resultado['motivo'] = exc.args[0] if exc.args else str(exc)
        except Exception as exc:  # noqa: BLE001
            _logger.exception('Error procesando fila %s de la hoja %s', numero_fila, hoja)
            resultado['motivo'] = _('Error inesperado: %s') % exc
        return resultado

    def _construir_vals(self, ramo: str, raw: dict) -> tuple:
        """Resuelve referencias y arma los vals de la póliza. Solo lectura.

        Lanza UserError con el motivo de rechazo si algo no se puede resolver.
        Las búsquedas de partner NO crean aquí (la creación ocurre en grabar).
        """
        nombre_poliza = self._txt(raw.get(COL_NUMERO_POLIZA[ramo]))
        if not nombre_poliza:
            raise UserError(_('Falta el número de póliza.'))
        agente = self._resolver_agente(raw.get('Clave de Agente'))
        producto = self._resolver_producto(self._txt(raw.get('Producto')), ramo)
        currency = self._resolver_moneda(raw.get('Moneda'))
        fecha_inicio = self._norm_fecha(raw.get('Fecha inicio Vigencia'))
        fecha_fin = self._norm_fecha(raw.get('Fecha Fin Vigencia'))
        if not fecha_inicio or not fecha_fin:
            raise UserError(_('Fechas de vigencia inválidas o ausentes.'))

        vals = {
            'name': nombre_poliza,
            'aseguradora_id': self.aseguradora_id.id,
            'producto_id': producto.id,
            'agente_id': agente.id,
            'currency_id': currency.id,
            'periodicidad': self._map_periodicidad(raw.get('Frecuencia de Pago')),
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'fecha_emision': self._norm_fecha(raw.get('Fecha emisión')),
            'plan': self._txt(raw.get('Plan')),
            'prima_anual': self._norm_monto(raw.get('Prima de Riesgo Anual')),
            'prima_fraccionada': self._norm_monto(raw.get('Prima Fraccionada')),
            'recargo_fijo': self._norm_monto(raw.get('Recargo Fijo')),
            'suma_asegurada': self._norm_monto(raw.get('Suma Asegurada')),
            'coberturas_adicionales': self._txt(raw.get('Coberturas Adicionales')),
            'estado': self._map_estado(raw.get('Estatus de Póliza')),
            'pagado_hasta_inicial': self._norm_fecha(raw.get('Pagado Hasta')),
            'pago_suspendido': self._es_suspendido(raw.get('Estatus de Pago')),
        }
        # Conducto (opcional: sin match no rechaza).
        conducto = self._resolver_conducto(raw.get('Conducto de Cobro'))
        if conducto:
            vals['conducto_id'] = conducto.id
        # Datos específicos por ramo.
        if ramo == 'gmm':
            vals.update({
                'bca_sub_ramo_codigo': self._txt(raw.get('Ramo / Sub ramo')),
                'nivel_hospitalario': self._txt(raw.get('Nivel Hospitalario')),
                'recargo_fraccionamiento': self._norm_monto(raw.get('Recargos (pago fraccionado)')),
                'iva': self._norm_monto(raw.get('IVA')),
                'deducible': self._norm_monto(raw.get('Deducible')),
                'coaseguro': self._norm_coaseguro(raw.get('Coaseguro')),
            })
            origen = self._buscar_poliza_origen(self._txt(raw.get('Póliza Original')))
            if origen:
                vals['poliza_origen_id'] = origen.id
        else:  # vida
            vals['tipo_cobertura'] = 'estandar'

        # Contratante (datos completos) y asegurado.
        vals['_contratante_data'] = self._datos_contratante(raw, ramo)
        vals['_asegurado_nombre'] = self._txt(raw.get('Nombre del Asegurado'))
        beneficiarios = self._construir_beneficiarios(raw, ramo)
        return vals, beneficiarios

    # ------------------------------------------------------------------ #
    # Creación / actualización
    # ------------------------------------------------------------------ #
    def _crear_poliza(self, vals: dict, beneficiarios: list, ramo: str):
        contratante_data = vals.pop('_contratante_data')
        asegurado_nombre = vals.pop('_asegurado_nombre')
        corte = vals.get('pagado_hasta_inicial')
        estado_destino = vals.pop('estado')  # se aplica tras generar el plan

        contratante = self._find_or_create_partner(contratante_data)
        vals['contratante_id'] = contratante.id
        if asegurado_nombre and asegurado_nombre != contratante_data.get('name'):
            asegurado = self._find_or_create_partner({'name': asegurado_nombre})
            vals['asegurado_id'] = asegurado.id

        poliza = self.env['bca.poliza'].create(vals)
        for b in beneficiarios:
            partner = self._find_or_create_partner({'name': b['nombre']})
            self.env['bca.poliza.beneficiario'].create({
                'poliza_id': poliza.id,
                'beneficiario_id': partner.id,
                'parentesco': b['parentesco'],
                'porcentaje': b['porcentaje'],
                'fecha_nacimiento': b['fecha_nacimiento'],
            })
        # Confirmación con corte: genera solo los recibos posteriores al
        # 'Pagado Hasta' declarado (decisión del usuario).
        if estado_destino == 'cancelada':
            poliza.estado = 'cancelada'
        else:
            poliza._validar_porcentaje_beneficiarios()
            poliza.estado = 'activa'
            poliza._generar_plan_pagos(desde=corte or poliza.fecha_inicio)
            if estado_destino == 'vencida':
                poliza.estado = 'vencida'
        return poliza

    def _actualizar_poliza(self, poliza, vals: dict):
        """Refresca solo datos informativos seguros en una póliza existente.

        No regenera el plan de recibos ni toca beneficiarios para no destruir
        historial de pagos (R-POL-05). Solo campos declarativos/contacto.
        """
        contratante_data = vals.pop('_contratante_data')
        vals.pop('_asegurado_nombre', None)
        campos_seguros = {
            k: vals[k] for k in (
                'suma_asegurada', 'pagado_hasta_inicial', 'pago_suspendido',
                'recargo_fijo', 'recargo_fraccionamiento', 'iva', 'deducible',
                'coaseguro', 'nivel_hospitalario', 'coberturas_adicionales',
            ) if k in vals
        }
        poliza.write(campos_seguros)
        # Refrescar datos de contacto del contratante.
        if poliza.contratante_id and contratante_data:
            datos = {k: v for k, v in contratante_data.items() if v and k != 'name'}
            if datos:
                poliza.contratante_id.write(datos)
        return poliza

    # ------------------------------------------------------------------ #
    # Resolvers (solo lectura; la creación de partners ocurre en grabar)
    # ------------------------------------------------------------------ #
    def _resolver_agente(self, clave_raw):
        clave = self._txt(clave_raw)
        if not clave:
            raise UserError(_('Falta la clave de agente.'))
        # La clave puede venir como número; normalizar a texto sin decimales.
        if isinstance(clave_raw, float) and clave_raw.is_integer():
            clave = str(int(clave_raw))
        Bridge = self.env['res.partner.agente.aseguradora']
        base = [('aseguradora_id', '=', self.aseguradora_id.id)]
        registro = Bridge.search(base + [('clave_agente', '=', clave)], limit=1)
        if not registro and clave.isdigit():
            # Tolerar ceros a la izquierda: el Excel pierde el padding al
            # tratar la clave como número (celda numérica), mientras la
            # aseguradora la registra con relleno (p. ej. '000019799' vs
            # '19799'). Se comparan por valor numérico en ambos sentidos.
            objetivo = clave.lstrip('0') or '0'
            registro = Bridge.search(base).filtered(
                lambda b: (b.clave_agente or '').strip().isdigit()
                and ((b.clave_agente or '').strip().lstrip('0') or '0') == objetivo
            )[:1]
        if not registro:
            raise UserError(_(
                'Clave de agente "%s" no registrada en la aseguradora.'
            ) % clave)
        return registro.agente_id

    def _resolver_producto(self, nombre: str, ramo: str):
        if not nombre:
            raise UserError(_('Falta el nombre del producto.'))
        dominio_base = [
            ('bca_es_producto_seguro', '=', True),
            ('bca_aseguradora_id', '=', self.aseguradora_id.id),
            ('bca_ramo', '=', ramo),
        ]
        Producto = self.env['product.template']
        producto = Producto.search(dominio_base + [('name', '=', nombre)], limit=1)
        if not producto:
            producto = Producto.search(dominio_base + [('name', 'ilike', nombre)], limit=1)
        if not producto:
            raise UserError(_('Producto "%s" no encontrado para el ramo %s.') % (nombre, ramo))
        return producto

    def _resolver_moneda(self, valor):
        codigo = (self._txt(valor) or 'MXN').upper()
        if codigo not in ('MXN', 'USD'):
            codigo = 'MXN'
        moneda = self.env.ref('base.%s' % codigo, raise_if_not_found=False)
        return moneda or self.env.company.currency_id

    def _resolver_conducto(self, valor):
        codigo = self._txt(valor)
        if not codigo:
            return False
        return self.env['bca.conducto'].search([
            ('aseguradora_id', '=', self.aseguradora_id.id),
            ('activo', '=', True),
            '|', ('codigo_archivo', '=ilike', codigo), ('name', '=ilike', codigo),
        ], limit=1)

    def _buscar_poliza_origen(self, nombre: str):
        if not nombre:
            return False
        return self.env['bca.poliza'].search([
            ('name', '=', nombre),
            ('aseguradora_id', '=', self.aseguradora_id.id),
        ], limit=1)

    def _resolver_poliza(self, folio: str):
        """Resuelve una póliza por folio dentro de la aseguradora del wizard.

        Devuelve el recordset (vacío si no existe). El folio es único por
        (name, aseguradora_id) — ver la restricción SQL de bca.poliza.
        """
        Poliza = self.env['bca.poliza']
        if not folio:
            return Poliza
        return Poliza.search([
            ('name', '=', folio),
            ('aseguradora_id', '=', self.aseguradora_id.id),
        ], limit=1)

    def _find_or_create_partner(self, datos: dict):
        """Busca una persona (contratante/asegurado/beneficiario) de forma
        ROLE-AGNÓSTICA y la crea si no existe. Idempotente.

        Un mismo res.partner puede acumular varios roles de póliza (contratante
        y asegurado a la vez) e incluso una posición de red (p. ej. agente), así
        que NO se segrega por rol ni se fija bca_tipo: los roles se derivan de
        las pólizas. Se excluyen las entidades de red (aseguradora/promotoria/
        holding), que nunca son la persona de una póliza. El contratante puede
        ser persona o empresa, por eso no se filtra por is_company.

        Orden de emparejamiento: RFC (vat) → nombre case-insensitive → nombre
        normalizado (sin acentos ni dobles espacios).
        """
        nombre = datos.get('name')
        if not nombre:
            raise UserError(_('Falta el nombre de un contacto requerido.'))
        Partner = self.env['res.partner']
        base = [('bca_tipo', 'not in', list(TIPOS_RED_EXCLUIDOS_POLIZA))]
        # 1) Por RFC: identificador más fuerte cuando viene en el layout.
        vat = datos.get('vat')
        if vat:
            partner = Partner.search(base + [('vat', '=ilike', vat)], limit=1)
            if partner:
                return partner
        # 2) Por nombre case-insensitive (captura mayúsculas/espacios laterales).
        partner = Partner.search(base + [('name', '=ilike', nombre)], limit=1)
        if partner:
            return partner
        # 3) Red final: normalización fuerte en Python (acentos, dobles espacios).
        #    Se acota la búsqueda por el primer token para no barrer toda la tabla.
        objetivo = Partner._bca_norm_nombre(nombre)
        primer_token = objetivo.split(' ')[0] if objetivo else ''
        if primer_token:
            candidatos = Partner.search(base + [('name', 'ilike', primer_token)])
            for cand in candidatos:
                if Partner._bca_norm_nombre(cand.name) == objetivo:
                    return cand
        return Partner.create(dict(datos))

    # ------------------------------------------------------------------ #
    # Construcción de datos auxiliares
    # ------------------------------------------------------------------ #
    def _datos_contratante(self, raw: dict, ramo: str) -> dict:
        ref_field = 'bca_ref_prima_medica' if ramo == 'gmm' else 'bca_ref_prima_basica_trad'
        ref_col = ('Referencia de cobro Prima (MEDICA)' if ramo == 'gmm'
                   else 'Referencia Prima Básica (TRAD)')
        datos = {
            'name': self._txt(raw.get('Nombre del Contratante')),
            'vat': self._txt(raw.get('R.F.C. Contratante')),
            'street': self._txt(raw.get('Calle y número')),
            'street2': self._txt(raw.get('Colonia')),
            'city': self._txt(raw.get('Población (Alcaldía o Municipio)')),
            'zip': self._txt(raw.get('C.P')),
            'phone': self._txt(raw.get('Teléfono o Celular')),
            'email': self._txt(raw.get('e-mail del Contratante')),
            'bca_fecha_nacimiento': self._norm_fecha(raw.get('Fecha de nacimiento')),
            'bca_estado_civil': self._map_simple(raw.get('Estado Civil'), ESTADO_CIVIL_MAP),
            'bca_genero': self._map_simple(raw.get('Género'), GENERO_MAP),
            ref_field: self._txt(raw.get(ref_col)),
        }
        if ramo == 'vida':
            datos.update({
                'bca_fondo_variable': self._txt(raw.get('Fondo Variable')),
                'bca_fondo_fijo': self._txt(raw.get('Fondo Fijo')),
                'bca_fondo_variable_ppr': self._txt(
                    raw.get('Fondo Variable Plan Personal de Retiro (PPR)')),
                'bca_fondo_fijo_ppr': self._txt(
                    raw.get('Fondo Fijo Plan Personal de Retiro (PPR)')),
                'bca_fondo_variable_cpea': self._txt(
                    raw.get('Fondo Variable Cuenta Personal Especial de Ahorro (CPEA)')),
                'bca_fondo_fijo_cpea': self._txt(
                    raw.get('Fondo Fijo Cuenta Especial de Ahorro (CPEA)')),
            })
        if not datos['name']:
            raise UserError(_('Falta el nombre del contratante.'))
        return {k: v for k, v in datos.items() if v}

    def _construir_beneficiarios(self, raw: dict, ramo: str) -> list:
        """VIDA: hasta 7 beneficiarios (con %). GMM: hasta 5 asegurados
        adicionales (con fecha de nacimiento). Ambos al mismo modelo."""
        beneficiarios: list[dict] = []
        if ramo == 'vida':
            for n in range(1, 8):
                nombre = self._txt(raw.get('Nombre del Beneficiario %d' % n))
                if not nombre:
                    continue
                beneficiarios.append({
                    'nombre': nombre,
                    'parentesco': self._map_simple(
                        raw.get('Parentesco %d' % n), PARENTESCO_MAP),
                    'porcentaje': self._norm_monto(
                        raw.get('%% al que tiene Derecho %d' % n)),
                    'fecha_nacimiento': False,
                })
        else:  # gmm
            for n in range(1, 6):
                nombre = self._txt(raw.get('Nombre del Asegurado %d' % n))
                if not nombre:
                    continue
                beneficiarios.append({
                    'nombre': nombre,
                    'parentesco': self._map_simple(
                        raw.get('Parentesco %d' % n), PARENTESCO_MAP),
                    'porcentaje': 0.0,
                    'fecha_nacimiento': self._norm_fecha(
                        raw.get('Fecha de nacimiento (Asegurado %d)' % n)),
                })
        return beneficiarios

    # ------------------------------------------------------------------ #
    # Normalizadores / mapeos
    # ------------------------------------------------------------------ #
    @staticmethod
    def _txt(valor) -> str:
        if valor is None:
            return ''
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        return str(valor).strip()

    def _norm_fecha(self, valor):
        if valor in (None, ''):
            return False
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        texto = str(valor).strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                continue
        return False

    def _norm_monto(self, valor) -> float:
        if valor in (None, ''):
            return 0.0
        if isinstance(valor, (int, float)):
            return float(valor)
        texto = str(valor).strip().replace(',', '').replace('$', '').replace('%', '')
        try:
            return float(texto) if texto else 0.0
        except ValueError:
            return 0.0

    def _norm_coaseguro(self, valor) -> float:
        """Coaseguro a fracción 0–1 (la póliza lo guarda como 0.05 = 5%).

        El layout puede traerlo como 5 (puntos %) o 0.05 (fracción). Se normaliza:
        valores > 1 se asumen en puntos porcentuales y se dividen entre 100.
        """
        val = self._norm_monto(valor)
        return val / 100.0 if val > 1 else val

    def _map_periodicidad(self, valor) -> str:
        texto = (self._txt(valor) or '').lower()
        for clave, destino in PERIODICIDAD_MAP.items():
            if clave in texto:
                return destino
        return 'anual'

    def _map_estado(self, valor) -> str:
        texto = (self._txt(valor) or '').lower()
        for clave, destino in ESTADO_POLIZA_MAP.items():
            if clave in texto:
                return destino
        return 'activa'

    def _es_suspendido(self, valor) -> bool:
        return 'suspend' in (self._txt(valor) or '').lower()

    @staticmethod
    def _map_simple(valor, mapa: dict):
        if valor in (None, ''):
            return False
        return mapa.get(str(valor).strip().lower(), False)

    # ------------------------------------------------------------------ #
    # Reporte / navegación
    # ------------------------------------------------------------------ #
    def _render_reporte(self, resultados: list, fase: str) -> str:
        ok = sum(1 for r in resultados if not r['motivo'])
        ko = sum(1 for r in resultados if r['motivo'])
        filas_html = []
        for r in resultados:
            color = '#b30000' if r['motivo'] else '#1a7f37'
            detalle = r['motivo'] or _('OK — %s') % (r['accion'] or '—')
            filas_html.append(
                '<tr><td>%s</td><td>%s</td><td>%s</td>'
                '<td style="color:%s">%s</td></tr>' % (
                    r['hoja'], r['fila'], r['poliza'], color, detalle)
            )
        return (
            '<p><strong>%s</strong>: %d correctas, %d con problema.</p>'
            '<table class="table table-sm">'
            '<thead><tr><th>Hoja</th><th>Fila</th><th>Póliza</th>'
            '<th>Resultado</th></tr></thead><tbody>%s</tbody></table>'
        ) % (fase.capitalize(), ok, ko, ''.join(filas_html))

    def _reabrir(self) -> dict:
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
