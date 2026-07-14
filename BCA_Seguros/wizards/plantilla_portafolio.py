"""Catálogo de columnas y constructor del libro Excel de la plantilla de
portafolio que consume el wizard ``bca.wizard.carga.portafolio``.

Este módulo es Python puro (solo depende de ``openpyxl``): NO importa Odoo, de
modo que lo reutilizan tanto la acción de descarga del wizard
(``wizards/carga_portafolio.py``) como el script de desarrollo
(``tools/generar_plantilla_portafolio.py``). El catálogo de encabezados vive
aquí en un único sitio (DRY).

Los nombres de los encabezados se mantienen sincronizados, carácter por
carácter, con las claves leídas en ``wizards/carga_portafolio.py``
(``raw.get(...)``). Cambiar un encabezado aquí sin cambiar el wizard — o
viceversa — descarta silenciosamente la columna. El test round-trip de
``tests/test_carga_portafolio.py`` protege contra esa deriva.

Las columnas marcadas ``REQUERIDO`` en la ayuda (comentario de celda del
encabezado) son las validadas por ``COLUMNAS_REQUERIDAS``: su ausencia aborta
la hoja completa. El resto son opcionales/informativas y una celda vacía es
válida.
"""

from __future__ import annotations

import io
import re
import unicodedata

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Catálogo de columnas (encabezado, ayuda de formato). El orden es estético;
# el wizard empareja por nombre de encabezado, no por posición.
# --------------------------------------------------------------------------- #
# Bloque común a ambos ramos salvo el número de póliza (encabezado distinto).
_COMUNES_IDENTIFICACION_VIDA = [
    ('Póliza', 'REQUERIDO · texto'),
]
_COMUNES_IDENTIFICACION_GMM = [
    ('Poliza actual', 'REQUERIDO · texto'),
]
_COMUNES_CABECERA = [
    ('Producto', 'REQUERIDO · nombre exacto del producto en el catálogo'),
    ('Plan', 'texto'),
    ('Clave de Agente', 'REQUERIDO · clave registrada en la aseguradora'),
    ('Conducto de Cobro', 'código o nombre del conducto (opcional)'),
    ('Moneda', 'REQUERIDO · MXN o USD'),
    ('Fecha emisión', 'fecha dd/mm/aaaa'),
    ('Fecha inicio Vigencia', 'REQUERIDO · fecha dd/mm/aaaa'),
    ('Fecha Fin Vigencia', 'REQUERIDO · fecha dd/mm/aaaa'),
    ('Frecuencia de Pago', 'REQUERIDO · Mensual/Trimestral/Semestral/Anual'),
]
_COMUNES_PRIMAS = [
    ('Prima de Riesgo Anual', 'importe'),
    ('Prima Fraccionada', 'importe por recibo'),
    ('Recargo Fijo', 'importe'),
    ('Suma Asegurada', 'importe'),
    ('Coberturas Adicionales', 'texto libre'),
    ('Estatus de Póliza', 'Vigente/Vencida/Cancelada'),
    ('Estatus de Pago', 'Al corriente/Vencido/Suspendido'),
    ('Pagado Hasta', 'fecha dd/mm/aaaa · ancla la generación de recibos'),
]
_CONTRATANTE = [
    ('Nombre del Contratante', 'REQUERIDO · razón social o nombre completo'),
    ('R.F.C. Contratante', 'RFC'),
    ('Fecha de nacimiento', 'fecha dd/mm/aaaa'),
    ('Estado Civil', 'Soltero/Casado/Divorciado/Viudo/Unión libre'),
    ('Género', 'Masculino/Femenino'),
    ('Calle y número', 'texto'),
    ('Colonia', 'texto'),
    ('Población (Alcaldía o Municipio)', 'texto'),
    ('C.P', 'código postal'),
    ('Teléfono o Celular', 'teléfono'),
    ('e-mail del Contratante', 'correo'),
]

# Vida: referencia de prima básica + asegurado titular + fondos.
_VIDA_ESPECIFICO = [
    ('Referencia Prima Básica (TRAD)', 'referencia de cobro (vida tradicional)'),
    ('Nombre del Asegurado', 'persona asegurada si difiere del contratante'),
    ('Fondo Variable', 'texto/importe'),
    ('Fondo Fijo', 'texto/importe'),
    ('Fondo Variable Plan Personal de Retiro (PPR)', 'texto/importe'),
    ('Fondo Fijo Plan Personal de Retiro (PPR)', 'texto/importe'),
    ('Fondo Variable Cuenta Personal Especial de Ahorro (CPEA)', 'texto/importe'),
    ('Fondo Fijo Cuenta Especial de Ahorro (CPEA)', 'texto/importe'),
]

# GMM: datos médicos antes de las primas comunes.
_GMM_MEDICO = [
    ('Ramo / Sub ramo', 'código de la aseguradora'),
    ('Nivel Hospitalario', 'texto'),
    ('Recargos (pago fraccionado)', 'importe'),
    ('IVA', 'importe'),
    ('Deducible', 'importe'),
    ('Coaseguro', 'porcentaje (10 = 10%) o fracción (0.10)'),
    ('Póliza Original', 'número de póliza origen (renovación/conversión)'),
]
_GMM_ESPECIFICO = [
    ('Referencia de cobro Prima (MEDICA)', 'referencia de cobro (gmm)'),
    ('Nombre del Asegurado', 'asegurado titular si difiere del contratante'),
]


# Beneficiarios (Vida) y asegurados adicionales/dependientes (GMM) viven en una
# hoja PROPIA en formato largo (una fila por persona), no como columnas anchas de
# las hojas de póliza: BCA los entrega en un documento distinto y una póliza puede
# tener hasta 10. La hoja se referencia contra la póliza por su folio (columna
# "Póliza") + la aseguradora seleccionada en el wizard. Ver B05.
_BENEFICIARIOS = [
    ('Póliza', 'REQUERIDO · folio de la póliza a la que pertenece'),
    ('Nombre del Beneficiario', 'REQUERIDO · nombre completo'),
    ('Parentesco', 'Cónyuge/Hijo/Padre/Madre/Hermano'),
    ('% al que tiene Derecho', 'Vida: porcentaje (por póliza debe sumar 100)'),
    ('Fecha de Nacimiento', 'GMM (dependiente): fecha dd/mm/aaaa'),
]


COLUMNAS_VIDA: list[tuple[str, str]] = (
    _COMUNES_IDENTIFICACION_VIDA
    + _COMUNES_CABECERA
    + _COMUNES_PRIMAS
    + _CONTRATANTE
    + _VIDA_ESPECIFICO
)

COLUMNAS_GMM: list[tuple[str, str]] = (
    _COMUNES_IDENTIFICACION_GMM
    + _COMUNES_CABECERA
    + _GMM_MEDICO
    + _COMUNES_PRIMAS
    + _CONTRATANTE
    + _GMM_ESPECIFICO
)

COLUMNAS_BENEFICIARIOS: list[tuple[str, str]] = _BENEFICIARIOS


# Coberturas adicionales por póliza (formato largo, una fila por cobertura),
# extraídas del archivo de "beneficios suplementarios" de MetLife. Reflejan sus
# 3 columnas para poder pegarlas directo. Se referencian contra la póliza por su
# folio ("Póliza", tolerante a ceros a la izquierda) + la aseguradora del wizard.
# El mapeo estructurado a la cobertura del catálogo se hace por la DESCRIPCIÓN
# (col C) vía COBERTURA_DESC_MAP; el CÓDIGO (col B) se conserva como texto.
_COBERTURAS = [
    ('Póliza', 'REQUERIDO · folio de la póliza a la que pertenece'),
    ('Cobertura Adicional', 'código de la cobertura en el archivo (ej. NVUAEP)'),
    ('Descripción Plan Suplementario',
     'REQUERIDO · descripción de la cobertura (ej. EXENCION DE PAGO DE PRIMAS INV.)'),
]

COLUMNAS_COBERTURAS: list[tuple[str, str]] = _COBERTURAS


# --------------------------------------------------------------------------- #
# Mapeo de coberturas MetLife: descripción (col C) → XML ID del valor de
# cobertura sembrado (product.attribute.value en data/coberturas_metlife.xml).
#
# El archivo trae CIENTOS de códigos (col B) que varían por producto/plan/moneda,
# pero la DESCRIPCIÓN se reduce a ~11 conceptos que ya coinciden 1:1 con los
# valores sembrados. Se mapea por descripción normalizada con coincidencia por
# FRAGMENTO (substring), de MÁS ESPECÍFICO A MÁS GENÉRICO: el orden importa
# porque "MUERTE ACCIDENTAL" es substring de "DOBLE/TRIPLE ... MUERTE ACCIDENTAL"
# y de "... PERDIDAS ORGANICAS". Los conceptos no sembrados (Cáncer/Diabetes/
# Invalidez de Met4U) quedan sin XML ID → sólo notas (no se pierde el dato).
# --------------------------------------------------------------------------- #
COBERTURA_DESC_MAP: list[tuple[str, str]] = [
    ('TRIPLE', 'val_ad_triple_muerte_accidental'),
    ('DOBLE', 'val_ad_doble_muerte_accidental'),
    ('PERDIDAS ORG', 'val_ad_muerte_accidental_po'),
    ('MUERTE ACCIDENTAL', 'val_ad_muerte_accidental'),
    ('EXENCION', 'val_ad_exencion_primas_invalidez'),
    ('PAGO ANTICIPADO', 'val_ad_pago_anticipado_invalidez'),
    ('GRAVES ENFERMEDADES', 'val_ad_graves_enfermedades'),
    ('GARANTIA PAGO PRIMAS FALLECIMIENTO', 'val_ad_garantia_fallecimiento'),
    ('GARANTIA PAGO PRIMAS POR INVALIDEZ', 'val_ad_garantia_primas_invalidez'),
]


def normalizar_texto_cobertura(txt) -> str:
    """Normaliza la descripción para el mapeo: quita artefactos de padding de
    ancho fijo (``Ê``/NBSP mal decodificados en Latin-1), acentos y puntuación;
    colapsa espacios y pasa a mayúsculas. Deja el token de producto al final,
    que es inocuo para la coincidencia por fragmento."""
    if not txt:
        return ''
    s = str(txt).replace('Ê', ' ').replace('\xa0', ' ')
    s = ''.join(
        c for c in unicodedata.normalize('NFKD', s)
        if not unicodedata.combining(c)
    )
    s = re.sub(r'[^A-Za-z0-9]+', ' ', s.upper())
    return re.sub(r'\s+', ' ', s).strip()


def normalizar_folio(txt) -> str:
    """Normaliza un folio para emparejar ignorando ceros a la izquierda
    (ej. ``'0008312115'`` y ``'8312115'`` colisionan)."""
    return str(txt or '').strip().lstrip('0')


def mapear_cobertura(descripcion) -> str | None:
    """Devuelve el XML ID (dentro de ``BCA_Seguros``) del valor de cobertura para
    la descripción dada, o ``None`` si no hay coincidencia (→ sólo notas)."""
    norm = normalizar_texto_cobertura(descripcion)
    if not norm:
        return None
    for fragmento, xmlid in COBERTURA_DESC_MAP:
        if fragmento in norm:
            return xmlid
    return None

# --------------------------------------------------------------------------- #
# Filas de ejemplo (indexadas por encabezado). Los encabezados ausentes se
# renderizan como celdas vacías.
# --------------------------------------------------------------------------- #
EJEMPLOS_VIDA = [
    {
        'Póliza': 'PV-0001', 'Producto': 'TempoLife', 'Plan': 'Tradicional',
        'Clave de Agente': 'A100', 'Conducto de Cobro': 'Domiciliación',
        'Moneda': 'MXN', 'Fecha emisión': '15/12/2024',
        'Fecha inicio Vigencia': '01/01/2025', 'Fecha Fin Vigencia': '01/01/2045',
        'Frecuencia de Pago': 'Mensual', 'Prima de Riesgo Anual': '24000.00',
        'Prima Fraccionada': '2000.00', 'Suma Asegurada': '1000000.00',
        'Estatus de Póliza': 'Vigente', 'Estatus de Pago': 'Al corriente',
        'Pagado Hasta': '30/06/2025', 'Nombre del Contratante': 'Juan Pérez García',
        'R.F.C. Contratante': 'PEGJ800101AAA', 'Fecha de nacimiento': '01/01/1980',
        'Estado Civil': 'Casado', 'Género': 'Masculino',
        'Calle y número': 'Av. Reforma 100', 'Colonia': 'Centro',
        'Población (Alcaldía o Municipio)': 'Cuauhtémoc', 'C.P': '06000',
        'Teléfono o Celular': '5512345678', 'e-mail del Contratante': 'juan@example.com',
        'Nombre del Asegurado': 'Juan Pérez García',
    },
    {
        'Póliza': 'PV-0002', 'Producto': 'TempoLife', 'Clave de Agente': 'A100',
        'Moneda': 'USD', 'Fecha inicio Vigencia': '01/03/2025',
        'Fecha Fin Vigencia': '01/03/2030', 'Frecuencia de Pago': 'Anual',
        'Prima de Riesgo Anual': '1200.00', 'Suma Asegurada': '50000.00',
        'Estatus de Póliza': 'Vigente', 'Estatus de Pago': 'Al corriente',
        'Nombre del Contratante': 'Ana Torres Ruiz',
        'R.F.C. Contratante': 'TORA900202BBB', 'Género': 'Femenino',
    },
]

EJEMPLOS_GMM = [
    {
        'Poliza actual': 'PG-0001', 'Producto': 'GMM Integral', 'Plan': 'Platino',
        'Clave de Agente': 'A100', 'Conducto de Cobro': 'Tarjeta',
        'Moneda': 'MXN', 'Fecha inicio Vigencia': '01/01/2025',
        'Fecha Fin Vigencia': '01/01/2026', 'Frecuencia de Pago': 'Anual',
        'Ramo / Sub ramo': 'GMM-IND', 'Nivel Hospitalario': 'A',
        'Prima de Riesgo Anual': '20000.00', 'IVA': '3200.00',
        'Deducible': '30000.00', 'Coaseguro': '10', 'Suma Asegurada': '5000000.00',
        'Estatus de Póliza': 'Vigente', 'Estatus de Pago': 'Al corriente',
        'Pagado Hasta': '01/01/2025', 'Nombre del Contratante': 'Ana López Díaz',
        'R.F.C. Contratante': 'LODA850303CCC', 'Género': 'Femenino',
    },
    {
        'Poliza actual': 'PG-0002', 'Producto': 'GMM Integral',
        'Clave de Agente': 'A100', 'Moneda': 'MXN',
        'Fecha inicio Vigencia': '15/02/2025', 'Fecha Fin Vigencia': '15/02/2026',
        'Frecuencia de Pago': 'Mensual', 'Prima de Riesgo Anual': '18000.00',
        'Deducible': '25000.00', 'Coaseguro': '0.05',
        'Estatus de Póliza': 'Vigente', 'Estatus de Pago': 'Al corriente',
        'Nombre del Contratante': 'Roberto Gómez Sánchez',
        'R.F.C. Contratante': 'GOSR780404DDD', 'Género': 'Masculino',
    },
]

# Formato largo: una fila por beneficiario/dependiente, referenciando la póliza
# por su folio. VIDA usa "% al que tiene Derecho" (suma 100 por póliza); GMM usa
# "Fecha de Nacimiento" (dependientes). PV-0001/PG-0001 refieren a los ejemplos
# de las hojas VIDA/GMM.
EJEMPLOS_BENEFICIARIOS = [
    {'Póliza': 'PV-0001', 'Nombre del Beneficiario': 'María Pérez',
     'Parentesco': 'Hija', '% al que tiene Derecho': '50'},
    {'Póliza': 'PV-0001', 'Nombre del Beneficiario': 'Pedro Pérez',
     'Parentesco': 'Hijo', '% al que tiene Derecho': '50'},
    {'Póliza': 'PV-0002', 'Nombre del Beneficiario': 'Luis Torres',
     'Parentesco': 'Cónyuge', '% al que tiene Derecho': '100'},
    {'Póliza': 'PG-0001', 'Nombre del Beneficiario': 'Carlos López',
     'Parentesco': 'Cónyuge', 'Fecha de Nacimiento': '15/05/1984'},
    {'Póliza': 'PG-0001', 'Nombre del Beneficiario': 'Sofía López',
     'Parentesco': 'Hija', 'Fecha de Nacimiento': '20/09/2012'},
]

# Formato largo: una fila por cobertura adicional, referenciando la póliza por su
# folio. PV-0001 refiere al ejemplo de la hoja VIDA. La descripción es la que se
# mapea al catálogo (el código es informativo/auditoría).
EJEMPLOS_COBERTURAS = [
    {'Póliza': 'PV-0001', 'Cobertura Adicional': 'NVUAEP',
     'Descripción Plan Suplementario': 'EXENCION DE PAGO DE PRIMAS INV.'},
    {'Póliza': 'PV-0001', 'Cobertura Adicional': 'NVUAPA',
     'Descripción Plan Suplementario': 'PAGO ANTICIPADO SA POR INVALIDEZ'},
    {'Póliza': 'PV-0001', 'Cobertura Adicional': 'NVUAGE',
     'Descripción Plan Suplementario': 'GRAVES ENFERMEDADES'},
]

# --------------------------------------------------------------------------- #
# Renderizado
# --------------------------------------------------------------------------- #
FILL_HEADER = PatternFill('solid', fgColor='2E75B6')
FONT_HEADER = Font(color='FFFFFF', bold=True, size=10)
WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _construir_hoja(wb, nombre: str, columnas: list, ejemplos: list) -> None:
    """Crea una hoja con la estructura que espera el wizard.

    Fila 1: encabezados (los que lee el wizard) · fila 2+: datos/ejemplos.
    La ayuda de formato de cada columna se adjunta como COMENTARIO de celda
    sobre su encabezado, para no ocupar una fila y no romper el layout
    "títulos en la primera fila, datos en la segunda".
    """
    ws = wb.create_sheet(nombre)
    headers = [c[0] for c in columnas]
    hints = [c[1] for c in columnas]

    for col, (header, hint) in enumerate(zip(headers, hints), start=1):
        celda = ws.cell(row=1, column=col, value=header)
        celda.font = FONT_HEADER
        celda.fill = FILL_HEADER
        celda.alignment = WRAP
        if hint:
            celda.comment = Comment(hint, 'Plantilla BCA')

    for fila_idx, ejemplo in enumerate(ejemplos, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(row=fila_idx, column=col, value=ejemplo.get(header, ''))

    for col, header in enumerate(headers, start=1):
        ancho = min(max(len(header) + 2, 14), 40)
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = 'B2'


def construir_workbook() -> openpyxl.Workbook:
    """Arma el libro con las hojas VIDA, GMM, BENEFICIARIOS y COBERTURAS. No
    escribe a disco."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _construir_hoja(wb, 'VIDA', COLUMNAS_VIDA, EJEMPLOS_VIDA)
    _construir_hoja(wb, 'GMM', COLUMNAS_GMM, EJEMPLOS_GMM)
    _construir_hoja(wb, 'BENEFICIARIOS', COLUMNAS_BENEFICIARIOS, EJEMPLOS_BENEFICIARIOS)
    _construir_hoja(wb, 'COBERTURAS', COLUMNAS_COBERTURAS, EJEMPLOS_COBERTURAS)
    return wb


def construir_workbook_bytes() -> bytes:
    """Devuelve los bytes del ``.xlsx`` de la plantilla (para adjuntar/descargar)."""
    buffer = io.BytesIO()
    construir_workbook().save(buffer)
    return buffer.getvalue()
